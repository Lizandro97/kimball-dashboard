"""Build multi-sheet Excel exports per dashboard module."""

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from api.services.chart_service import filter_data, load_data


def _style_header(ws, ncols):
    hfont = Font(name="Inter", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin = Side(style="thin", color="E2E8F0")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)


def _write_rows(ws, rows):
    thin = Side(style="thin", color="E2E8F0")
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = Font(name="Inter", size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(bottom=thin)
    for c in range(1, len(rows[0]) + 1 if rows else 1):
        ws.column_dimensions[get_column_letter(c)].width = 18


def _add_sheet(wb, title: str, headers: list, rows: list):
    ws = wb.create_sheet(title=title)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, len(headers))
    _write_rows(ws, rows)
    return ws


def build_overview_excel(region: str, year: str, segment: str) -> io.BytesIO:
    df_global = load_data()
    df = filter_data(df_global, region, year, segment)

    sales = float(df["sales"].sum())
    profit = float(df["profit"].sum())
    orders = int(df["order_id"].nunique())
    customers = int(df["customer_name"].nunique())
    margin = round(profit / sales * 100, 1) if sales else 0

    wb = Workbook()
    wb.remove(wb.active)

    _add_sheet(wb, "KPIs", ["Indicador", "Valor"], [
        ["Ingresos Totales", round(sales, 2)],
        ["Utilidad Total", round(profit, 2)],
        ["Ordenes", orders],
        ["Clientes", customers],
        ["Margen %", margin],
        ["Unidades Vendidas", int(df["quantity"].sum())],
    ])

    top5 = (
        df.groupby(["product_name", "category"])
        .agg(revenue=("sales", "sum"), profit=("profit", "sum"))
        .assign(margin=lambda x: (x["profit"] / x["revenue"] * 100).round(1))
        .sort_values("revenue", ascending=False).head(5).reset_index()
    )
    _add_sheet(wb, "Top 5 Productos",
        ["Producto", "Categoria", "Ingresos", "Utilidad", "Margen %"],
        [[r["product_name"], r["category"], round(r["revenue"], 2),
          round(r["profit"], 2), r["margin"]] for _, r in top5.iterrows()]
    )

    pareto = (
        df.groupby("product_name")["sales"].sum()
        .sort_values(ascending=False).reset_index()
    )
    pareto["cum_pct"] = pareto["sales"].cumsum() / pareto["sales"].sum() * 100
    pareto["product_pct"] = (pareto.index + 1) / len(pareto) * 100
    _add_sheet(wb, "Pareto 80-20",
        ["#", "Producto", "Ingresos", "% Acumulado", "% Productos"],
        [[i + 1, r["product_name"], round(r["sales"], 2),
          round(r["cum_pct"], 1), round(r["product_pct"], 1)]
         for i, (_, r) in enumerate(pareto.head(20).iterrows())]
    )

    monthly = df.set_index("order_date").resample("ME")["sales"].sum().reset_index()
    monthly["mom_pct"] = monthly["sales"].pct_change() * 100
    _add_sheet(wb, "Tendencia Mensual",
        ["Mes", "Ventas", "Var % Mensual"],
        [[r["order_date"].strftime("%Y-%m"), round(r["sales"], 2),
          round(r["mom_pct"], 1) if pd.notna(r["mom_pct"]) else ""]
         for _, r in monthly.iterrows()]
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_sales_excel(region: str, year: str, segment: str) -> io.BytesIO:
    df_global = load_data()
    df = filter_data(df_global, region, year, segment)

    wb = Workbook()
    wb.remove(wb.active)

    top = (
        df.groupby(["product_name", "category", "sub_category"])
        .agg(ventas=("sales", "sum"), utilidad=("profit", "sum"))
        .assign(margen=lambda x: (x["utilidad"] / x["ventas"] * 100).round(1))
        .sort_values("ventas", ascending=False).head(10).reset_index()
    )
    _add_sheet(wb, "Top 10 Productos",
        ["#", "Producto", "Categoria", "Subcategoria", "Ventas", "Utilidad", "Margen %"],
        [[i + 1, r["product_name"], r["category"], r["sub_category"],
          round(r["ventas"], 2), round(r["utilidad"], 2), r["margen"]]
         for i, (_, r) in enumerate(top.iterrows())]
    )

    by_cat = df.groupby("category")["sales"].sum().reset_index().sort_values("sales", ascending=False)
    _add_sheet(wb, "Ventas por Categoria",
        ["Categoria", "Ventas"],
        [[r["category"], round(r["sales"], 2)] for _, r in by_cat.iterrows()]
    )

    by_reg = df.groupby("region")["sales"].sum().reset_index().sort_values("sales", ascending=False)
    _add_sheet(wb, "Ventas por Region",
        ["Region", "Ventas"],
        [[r["region"], round(r["sales"], 2)] for _, r in by_reg.iterrows()]
    )

    monthly = df.set_index("order_date").resample("ME")["sales"].sum().reset_index()
    _add_sheet(wb, "Tendencia Mensual",
        ["Mes", "Ventas"],
        [[r["order_date"].strftime("%Y-%m"), round(r["sales"], 2)] for _, r in monthly.iterrows()]
    )

    by_sub = df.groupby("sub_category")["sales"].sum().reset_index().sort_values("sales", ascending=False).head(15)
    _add_sheet(wb, "Top Subcategorias",
        ["Subcategoria", "Ventas"],
        [[r["sub_category"], round(r["sales"], 2)] for _, r in by_sub.iterrows()]
    )

    by_city = df.groupby(["city", "state"])["sales"].sum().reset_index().sort_values("sales", ascending=False).head(15)
    _add_sheet(wb, "Top Ciudades",
        ["Ciudad", "Estado", "Ventas"],
        [[r["city"], r["state"], round(r["sales"], 2)] for _, r in by_city.iterrows()]
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_profitability_excel(region: str, year: str, segment: str) -> io.BytesIO:
    df_global = load_data()
    df = filter_data(df_global, region, year, segment)

    wb = Workbook()
    wb.remove(wb.active)

    def tier(d):
        if d == 0: return "Sin desc."
        if d <= 0.2: return "≤20%"
        if d <= 0.5: return "≤50%"
        return ">50%"
    dfc = df.copy()
    dfc["tier"] = dfc["discount"].apply(tier)
    summary = dfc.groupby("tier", observed=True).agg(
        transacciones=("sales", "count"), utilidad=("profit", "sum")
    ).reset_index()
    _add_sheet(wb, "Niveles de Descuento",
        ["Nivel", "Transacciones", "Utilidad"],
        [[r["tier"], int(r["transacciones"]), round(r["utilidad"], 2)] for _, r in summary.iterrows()]
    )

    discount_loss = float(df[df["discount"] > 0.5]["profit"].sum())
    discount_pct = round(len(df[df["discount"] > 0.5]) / len(df) * 100, 1) if len(df) else 0
    _add_sheet(wb, "Perdida por Descuento",
        ["Indicador", "Valor"],
        [
            ["Perdida desc. agresivos", round(discount_loss, 2)],
            ["% transacciones >50% desc.", discount_pct],
        ]
    )

    rc = df.groupby(["region", "category"])["profit"].sum().reset_index().sort_values("profit", ascending=False)
    _add_sheet(wb, "Utilidad Region-Categoria",
        ["Region", "Categoria", "Utilidad"],
        [[r["region"], r["category"], round(r["profit"], 2)] for _, r in rc.iterrows()]
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_customers_excel(region: str, year: str, segment: str) -> io.BytesIO:
    df_global = load_data()
    df = filter_data(df_global, region, year, segment)

    wb = Workbook()
    wb.remove(wb.active)

    segments = (
        df.groupby("segment")
        .agg(
            ingresos=("sales", "sum"), ordenes=("order_id", "nunique"),
            clientes=("customer_name", "nunique"),
            total_profit=("profit", "sum"), avg_discount=("discount", "mean"),
        ).reset_index()
    )
    seg_rows = []
    for _, r in segments.iterrows():
        avg_order = r["ingresos"] / r["ordenes"] if r["ordenes"] else 0
        margin = round(r["total_profit"] / r["ingresos"] * 100, 1) if r["ingresos"] else 0
        seg_rows.append([
            r["segment"], int(r["clientes"]), round(r["ingresos"], 2),
            int(r["ordenes"]), round(avg_order, 2), margin,
            round(float(r["avg_discount"]) * 100, 1),
        ])
    _add_sheet(wb, "Segmentos",
        ["Segmento", "Clientes", "Ingresos", "Ordenes",
         "Ticket Promedio", "Margen %", "Descuento Prom %"],
        seg_rows
    )

    cust = (
        df.groupby(["customer_name", "segment"])
        .agg(revenue=("sales", "sum"), orders=("order_id", "nunique"),
             profit=("profit", "sum"))
        .reset_index()
    )
    cust["ticket"] = cust["revenue"] / cust["orders"]
    cust["margin"] = (cust["profit"] / cust["revenue"] * 100).round(1)
    top = cust.sort_values("revenue", ascending=False).head(10)
    _add_sheet(wb, "Top 10 Clientes",
        ["#", "Cliente", "Segmento", "Ingresos", "Ordenes",
         "Ticket Promedio", "Margen %"],
        [[i + 1, r["customer_name"], r["segment"], round(r["revenue"], 2),
          int(r["orders"]), round(r["ticket"], 2), r["margin"]]
         for i, (_, r) in enumerate(top.iterrows())]
    )

    freq = df.groupby("customer_name")["order_id"].nunique().reset_index()
    freq.columns = ["customer", "orders"]
    bins = [(1, 1, "1"), (2, 3, "2-3"), (4, 5, "4-5"), (6, 10, "6-10"), (11, 50, "11+")]
    freq_rows = []
    for lo, hi, label in bins:
        freq_rows.append([label, int(freq[(freq["orders"] >= lo) & (freq["orders"] <= hi)].shape[0])])
    _add_sheet(wb, "Frecuencia de Compra",
        ["Rango de Ordenes", "Clientes"],
        freq_rows
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_shipping_excel(region: str, year: str, segment: str) -> io.BytesIO:
    df_global = load_data()
    df = filter_data(df_global, region, year, segment)

    wb = Workbook()
    wb.remove(wb.active)

    d = df[["order_date", "ship_date", "ship_mode", "region"]].dropna().copy()
    d["delivery_days"] = (d["ship_date"] - d["order_date"]).dt.days

    avg_days = round(d["delivery_days"].mean(), 1)
    on_time = round((d["delivery_days"] <= 5).mean() * 100, 1)
    _add_sheet(wb, "Estadisticas Generales",
        ["Indicador", "Valor"],
        [
            ["Dias de entrega promedio", avg_days],
            ["% a tiempo (≤5 dias)", on_time],
            ["Total ordenes enviadas", int(d["ship_date"].count())],
        ]
    )

    by_region = d.groupby("region").agg(
        avg_days=("delivery_days", "mean"),
        on_time=("delivery_days", lambda x: round((x <= 5).mean() * 100, 1)),
        orders=("delivery_days", "count"),
    ).round(1).reset_index()
    _add_sheet(wb, "KPIs por Region",
        ["Region", "Dias Promedio", "% A Tiempo", "Ordenes"],
        [[r["region"], r["avg_days"], r["on_time"], int(r["orders"])]
         for _, r in by_region.iterrows()]
    )

    mode_data = d.groupby("ship_mode").agg(
        orders=("delivery_days", "count"),
        avg_days=("delivery_days", "mean"),
    ).round(1).reset_index()
    # Also add revenue/profit per mode
    mode_fin = df.groupby("ship_mode").agg(
        revenue=("sales", "sum"), profit=("profit", "sum"),
        orders=("order_id", "nunique"),
    ).reset_index()
    mode_fin["margin"] = (mode_fin["profit"] / mode_fin["revenue"] * 100).round(1)
    _add_sheet(wb, "Por Modo de Envio",
        ["Modo", "Ordenes", "Ingresos", "Utilidad", "Margen %"],
        [[r["ship_mode"], int(r["orders"]), round(r["revenue"], 2),
          round(r["profit"], 2), r["margin"]]
         for _, r in mode_fin.iterrows()]
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


EXPORT_BUILDERS = {
    "overview": build_overview_excel,
    "sales": build_sales_excel,
    "profitability": build_profitability_excel,
    "customers": build_customers_excel,
    "shipping": build_shipping_excel,
}